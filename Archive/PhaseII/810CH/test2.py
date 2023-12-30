''' Test 2 - https://chat.openai.com/share/d8e1b2a0-8413-4635-83e3-8ef3f7b65432 - convo link '''
# Rated 8.63/10
import pandas as pd
import os
import openpyxl
import numpy as np

# Define the output file name
output_file = 'trabajo_cotidiano_semestral.xlsx'

# Create an empty DataFrame to hold all the data
dataframe = pd.DataFrame()

# Keep track of the most recently modified file and its modification time
most_recent_file = None
most_recent_time = 0

# Keep track of the sum of all values in cell 'C3' (which is at index (2, 2))
total_puntos = 0

# Loop over every file in the directory
for file in os.listdir():
    # Check if the file is an xlsm file
    if file.endswith('.xlsm'):
        # Check the file modification time
        mod_time = os.path.getmtime(file)
        if mod_time > most_recent_time:
            most_recent_time = mod_time
            most_recent_file = file

        # Load the workbook
        workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)

        # Check each sheet
        for sheet in workbook.sheetnames:
            # If the sheet name contains 'dashboard' (case-insensitive)
            if 'dashboard' in sheet.lower():
                # Use openpyxl to get the 'C3' value directly
                ws = workbook[sheet]
                total_puntos += ws['C3'].value if ws['C3'].value else 0

                # Load the data into a DataFrame using pandas
                df = pd.read_excel(file, sheet_name=sheet)

                # Extract data from row 8 to 37 of the third column (column 'C' in Excel)
                column_data = df.iloc[7:37, 2]

                # Add the data to the final DataFrame
                dataframe[file] = column_data.reset_index(drop=True)

# Add a 'Total' column as the first column of the DataFrame
dataframe.insert(0, 'Total', dataframe.sum(axis=1))

# If there was at least one .xlsm file
if most_recent_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(most_recent_file, data_only=True)

    # Find the dashboard sheet
    for sheet in workbook.sheetnames:
        if 'dashboard' in sheet.lower():
            # Load the data into a DataFrame
            df = pd.read_excel(most_recent_file, sheet_name=sheet)

            # Extract data from row 8 to 37 of the second column (column 'B' in Excel)
            column_b_data = df.iloc[7:37, 1]

            # Insert this data to the left of the 'Total' column
            dataframe.insert(0, 'Estudiantes', column_b_data.reset_index(drop=True))

# Write the final DataFrame to the output file
dataframe.to_excel(output_file, index=False)

# Multiply the total puntos by 3
total_puntos *= 3

# Load the final workbook
workbook = openpyxl.load_workbook(output_file)

# Select the first sheet
sheet = workbook.active

# Write "Total Puntos:" to cell 'C35'
sheet['C35'] = "Total Puntos:"

# Write the total puntos to cell 'D35'
sheet['D35'] = total_puntos

# Save the workbook
workbook.save(output_file)

# Now let's add the 'Nota' and 'Porcentaje' columns to the dataframe and save it again

# Load the dataframe
df = pd.read_excel(output_file)

# Create a new column 'Nota' after 'Estudiantes' and fill it with 'Total' * (100 / total_puntos)
df.insert(1, 'Nota', df['Total'] * (100 / total_puntos))

# Round 'Nota' based on the specified rules
df['Nota'] = np.where((df['Nota'] % 1) < 0.5, np.floor(df['Nota']), np.ceil(df['Nota']))
df['Nota'] = df['Nota'].apply(lambda x: round(x, 2))

# Create a new column 'Porcentaje' after 'Nota' and fill it with 'Nota' * 0.6
df.insert(2, 'Porcentaje', df['Nota'] * 0.6)

# Round 'Porcentaje' based on the specified rules
df['Porcentaje'] = np.where((df['Porcentaje'] % 1) < 0.5, np.floor(df['Porcentaje']), np.ceil(df['Porcentaje']))
df['Porcentaje'] = df['Porcentaje'].apply(lambda x: round(x, 2))

# Delete column 12 from the DataFrame
df = df.drop(df.columns[11], axis=1)

# Rename 'Total' to 'Puntos Obtenidos'
df = df.rename(columns={'Total': 'Puntos Obtenidos'})

# Reorder the columns to have 'Porcentaje' as the third column
columns = df.columns.tolist()
columns = columns[:2] + ['Porcentaje'] + columns[2:]
df = df[columns]

# Save the updated dataframe back to the output file
df.to_excel(output_file, index=False)